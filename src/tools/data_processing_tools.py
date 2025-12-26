"""Data processing tools for CSV, JSON, YAML, XML, Excel, TOML"""

import csv
import json
import logging
from pathlib import Path
from typing import Dict, List, Optional, Any

from .security import validate_path

logger = logging.getLogger('data_processing_tools')

# Try to import optional dependencies
try:
    import yaml
    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False

try:
    import xml.etree.ElementTree as ET
    XML_AVAILABLE = True
except ImportError:
    XML_AVAILABLE = False

try:
    import openpyxl
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

try:
    import toml
    TOML_AVAILABLE = True
except ImportError:
    TOML_AVAILABLE = False


def _error_response(error: str, error_type: str) -> Dict[str, Any]:
    """Create standardized error response"""
    return {
        'success': False,
        'result': None,
        'error': error,
        'error_type': error_type
    }


def _success_response(result: Any) -> Dict[str, Any]:
    """Create standardized success response"""
    return {
        'success': True,
        'result': result,
        'error': None,
        'error_type': None
    }


def read_csv(file_path: str, delimiter: str = ',', encoding: str = 'utf-8') -> Dict[str, Any]:
    """
    Read CSV file
    
    Args:
        file_path: Path to CSV file
        delimiter: CSV delimiter
        encoding: File encoding
        
    Returns:
        Dict with success, result (list of dicts), error, error_type
    """
    try:
        is_valid, error_msg = validate_path(file_path, operation='read')
        if not is_valid:
            return _error_response(error_msg, 'ValidationError')
        
        path_obj = Path(file_path)
        rows = []
        
        with open(path_obj, 'r', encoding=encoding, newline='') as f:
            reader = csv.DictReader(f, delimiter=delimiter)
            rows = list(reader)
        
        return _success_response(rows)
    except Exception as e:
        logger.error(f"Error reading CSV {file_path}: {e}", exc_info=True)
        return _error_response(str(e), type(e).__name__)


def write_csv(file_path: str, data: List[Dict], headers: Optional[List[str]] = None, delimiter: str = ',') -> Dict[str, Any]:
    """
    Write CSV file
    
    Args:
        file_path: Path to CSV file
        data: List of dicts to write
        headers: Column headers (default: keys from first dict)
        delimiter: CSV delimiter
        
    Returns:
        Dict with success, result (file path), error, error_type
    """
    try:
        is_valid, error_msg = validate_path(file_path, operation='write', allow_absolute=True)
        if not is_valid:
            return _error_response(error_msg, 'ValidationError')
        
        path_obj = Path(file_path)
        path_obj.parent.mkdir(parents=True, exist_ok=True)
        
        if not data:
            return _error_response("No data to write", 'ValueError')
        
        fieldnames = headers or list(data[0].keys())
        
        with open(path_obj, 'w', encoding='utf-8', newline='') as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames, delimiter=delimiter)
            writer.writeheader()
            writer.writerows(data)
        
        return _success_response(str(path_obj))
    except Exception as e:
        logger.error(f"Error writing CSV {file_path}: {e}", exc_info=True)
        return _error_response(str(e), type(e).__name__)


def read_json(file_path: str, encoding: str = 'utf-8') -> Dict[str, Any]:
    """
    Read JSON file
    
    Args:
        file_path: Path to JSON file
        encoding: File encoding
        
    Returns:
        Dict with success, result (parsed JSON), error, error_type
    """
    try:
        is_valid, error_msg = validate_path(file_path, operation='read')
        if not is_valid:
            return _error_response(error_msg, 'ValidationError')
        
        path_obj = Path(file_path)
        
        with open(path_obj, 'r', encoding=encoding) as f:
            data = json.load(f)
        
        return _success_response(data)
    except json.JSONDecodeError as e:
        logger.error(f"JSON decode error in {file_path}: {e}", exc_info=True)
        return _error_response(f"Invalid JSON: {str(e)}", 'JSONDecodeError')
    except Exception as e:
        logger.error(f"Error reading JSON {file_path}: {e}", exc_info=True)
        return _error_response(str(e), type(e).__name__)


def write_json(file_path: str, data: Any, indent: int = 2, encoding: str = 'utf-8') -> Dict[str, Any]:
    """
    Write JSON file
    
    Args:
        file_path: Path to JSON file
        data: Data to write
        indent: JSON indentation
        encoding: File encoding
        
    Returns:
        Dict with success, result (file path), error, error_type
    """
    try:
        is_valid, error_msg = validate_path(file_path, operation='write', allow_absolute=True)
        if not is_valid:
            return _error_response(error_msg, 'ValidationError')
        
        path_obj = Path(file_path)
        path_obj.parent.mkdir(parents=True, exist_ok=True)
        
        with open(path_obj, 'w', encoding=encoding) as f:
            json.dump(data, f, indent=indent, ensure_ascii=False)
        
        return _success_response(str(path_obj))
    except Exception as e:
        logger.error(f"Error writing JSON {file_path}: {e}", exc_info=True)
        return _error_response(str(e), type(e).__name__)


def read_yaml(file_path: str, encoding: str = 'utf-8') -> Dict[str, Any]:
    """
    Read YAML file
    
    Args:
        file_path: Path to YAML file
        encoding: File encoding
        
    Returns:
        Dict with success, result (parsed YAML), error, error_type
    """
    if not YAML_AVAILABLE:
        return _error_response("PyYAML not available. Install with: pip install pyyaml", 'ImportError')
    
    try:
        is_valid, error_msg = validate_path(file_path, operation='read')
        if not is_valid:
            return _error_response(error_msg, 'ValidationError')
        
        path_obj = Path(file_path)
        
        with open(path_obj, 'r', encoding=encoding) as f:
            data = yaml.safe_load(f)
        
        return _success_response(data)
    except Exception as e:
        logger.error(f"Error reading YAML {file_path}: {e}", exc_info=True)
        return _error_response(str(e), type(e).__name__)


def write_yaml(file_path: str, data: Any, encoding: str = 'utf-8') -> Dict[str, Any]:
    """
    Write YAML file
    
    Args:
        file_path: Path to YAML file
        data: Data to write
        encoding: File encoding
        
    Returns:
        Dict with success, result (file path), error, error_type
    """
    if not YAML_AVAILABLE:
        return _error_response("PyYAML not available. Install with: pip install pyyaml", 'ImportError')
    
    try:
        is_valid, error_msg = validate_path(file_path, operation='write', allow_absolute=True)
        if not is_valid:
            return _error_response(error_msg, 'ValidationError')
        
        path_obj = Path(file_path)
        path_obj.parent.mkdir(parents=True, exist_ok=True)
        
        with open(path_obj, 'w', encoding=encoding) as f:
            yaml.dump(data, f, default_flow_style=False, allow_unicode=True)
        
        return _success_response(str(path_obj))
    except Exception as e:
        logger.error(f"Error writing YAML {file_path}: {e}", exc_info=True)
        return _error_response(str(e), type(e).__name__)


def read_xml(file_path: str, encoding: str = 'utf-8') -> Dict[str, Any]:
    """
    Read XML file
    
    Args:
        file_path: Path to XML file
        encoding: File encoding
        
    Returns:
        Dict with success, result (parsed XML as dict), error, error_type
    """
    if not XML_AVAILABLE:
        return _error_response("XML support not available", 'ImportError')
    
    try:
        is_valid, error_msg = validate_path(file_path, operation='read')
        if not is_valid:
            return _error_response(error_msg, 'ValidationError')
        
        path_obj = Path(file_path)
        tree = ET.parse(path_obj)
        root = tree.getroot()
        
        def xml_to_dict(element):
            result = {}
            if element.text and element.text.strip():
                result['text'] = element.text.strip()
            result['tag'] = element.tag
            result['attrib'] = element.attrib
            children = [xml_to_dict(child) for child in element]
            if children:
                result['children'] = children
            return result
        
        data = xml_to_dict(root)
        return _success_response(data)
    except ET.ParseError as e:
        logger.error(f"XML parse error in {file_path}: {e}", exc_info=True)
        return _error_response(f"Invalid XML: {str(e)}", 'ParseError')
    except Exception as e:
        logger.error(f"Error reading XML {file_path}: {e}", exc_info=True)
        return _error_response(str(e), type(e).__name__)


def write_xml(file_path: str, data: Dict, encoding: str = 'utf-8') -> Dict[str, Any]:
    """
    Write XML file
    
    Args:
        file_path: Path to XML file
        data: Data dict with 'tag', 'attrib', 'text', 'children'
        encoding: File encoding
        
    Returns:
        Dict with success, result (file path), error, error_type
    """
    if not XML_AVAILABLE:
        return _error_response("XML support not available", 'ImportError')
    
    try:
        is_valid, error_msg = validate_path(file_path, operation='write', allow_absolute=True)
        if not is_valid:
            return _error_response(error_msg, 'ValidationError')
        
        path_obj = Path(file_path)
        path_obj.parent.mkdir(parents=True, exist_ok=True)
        
        def dict_to_xml(parent, data):
            element = ET.SubElement(parent, data.get('tag', 'element'))
            if 'attrib' in data:
                element.attrib.update(data['attrib'])
            if 'text' in data:
                element.text = data['text']
            if 'children' in data:
                for child_data in data['children']:
                    dict_to_xml(element, child_data)
            return element
        
        root = ET.Element('root')
        dict_to_xml(root, data)
        tree = ET.ElementTree(root)
        tree.write(path_obj, encoding=encoding, xml_declaration=True)
        
        return _success_response(str(path_obj))
    except Exception as e:
        logger.error(f"Error writing XML {file_path}: {e}", exc_info=True)
        return _error_response(str(e), type(e).__name__)


def read_excel(file_path: str, sheet_name: Optional[str] = None) -> Dict[str, Any]:
    """
    Read Excel file
    
    Args:
        file_path: Path to Excel file
        sheet_name: Sheet name (default: first sheet)
        
    Returns:
        Dict with success, result (dict of sheet data), error, error_type
    """
    if not EXCEL_AVAILABLE:
        return _error_response("openpyxl not available. Install with: pip install openpyxl", 'ImportError')
    
    try:
        is_valid, error_msg = validate_path(file_path, operation='read')
        if not is_valid:
            return _error_response(error_msg, 'ValidationError')
        
        path_obj = Path(file_path)
        wb = openpyxl.load_workbook(path_obj, data_only=True)
        
        result = {}
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return _error_response(f"Sheet '{sheet_name}' not found", 'ValueError')
            sheets = [sheet_name]
        else:
            sheets = wb.sheetnames
        
        for sheet in sheets:
            ws = wb[sheet]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append(list(row))
            result[sheet] = rows
        
        return _success_response(result)
    except Exception as e:
        logger.error(f"Error reading Excel {file_path}: {e}", exc_info=True)
        return _error_response(str(e), type(e).__name__)


def write_excel(file_path: str, data: Dict[str, List], sheet_name: str = 'Sheet1') -> Dict[str, Any]:
    """
    Write Excel file
    
    Args:
        file_path: Path to Excel file
        data: Dict mapping sheet names to list of rows
        sheet_name: Default sheet name
        
    Returns:
        Dict with success, result (file path), error, error_type
    """
    if not EXCEL_AVAILABLE:
        return _error_response("openpyxl not available. Install with: pip install openpyxl", 'ImportError')
    
    try:
        is_valid, error_msg = validate_path(file_path, operation='write', allow_absolute=True)
        if not is_valid:
            return _error_response(error_msg, 'ValidationError')
        
        path_obj = Path(file_path)
        path_obj.parent.mkdir(parents=True, exist_ok=True)
        
        wb = openpyxl.Workbook()
        
        # Remove default sheet if we have data
        if data:
            wb.remove(wb.active)
        
        for sheet, rows in data.items():
            ws = wb.create_sheet(title=sheet)
            for row in rows:
                ws.append(row)
        
        if not data:
            ws = wb.active
            ws.title = sheet_name
        
        wb.save(path_obj)
        
        return _success_response(str(path_obj))
    except Exception as e:
        logger.error(f"Error writing Excel {file_path}: {e}", exc_info=True)
        return _error_response(str(e), type(e).__name__)


def read_toml(file_path: str, encoding: str = 'utf-8') -> Dict[str, Any]:
    """
    Read TOML file
    
    Args:
        file_path: Path to TOML file
        encoding: File encoding
        
    Returns:
        Dict with success, result (parsed TOML), error, error_type
    """
    if not TOML_AVAILABLE:
        return _error_response("toml library not available. Install with: pip install toml", 'ImportError')
    
    try:
        is_valid, error_msg = validate_path(file_path, operation='read')
        if not is_valid:
            return _error_response(error_msg, 'ValidationError')
        
        path_obj = Path(file_path)
        
        with open(path_obj, 'r', encoding=encoding) as f:
            data = toml.load(f)
        
        return _success_response(data)
    except Exception as e:
        logger.error(f"Error reading TOML {file_path}: {e}", exc_info=True)
        return _error_response(str(e), type(e).__name__)


def write_toml(file_path: str, data: Dict, encoding: str = 'utf-8') -> Dict[str, Any]:
    """
    Write TOML file
    
    Args:
        file_path: Path to TOML file
        data: Data to write
        encoding: File encoding
        
    Returns:
        Dict with success, result (file path), error, error_type
    """
    if not TOML_AVAILABLE:
        return _error_response("toml library not available. Install with: pip install toml", 'ImportError')
    
    try:
        is_valid, error_msg = validate_path(file_path, operation='write', allow_absolute=True)
        if not is_valid:
            return _error_response(error_msg, 'ValidationError')
        
        path_obj = Path(file_path)
        path_obj.parent.mkdir(parents=True, exist_ok=True)
        
        with open(path_obj, 'w', encoding=encoding) as f:
            toml.dump(data, f)
        
        return _success_response(str(path_obj))
    except Exception as e:
        logger.error(f"Error writing TOML {file_path}: {e}", exc_info=True)
        return _error_response(str(e), type(e).__name__)


def parse_json_string(json_string: str) -> Dict[str, Any]:
    """
    Parse JSON string
    
    Args:
        json_string: JSON string to parse
        
    Returns:
        Dict with success, result (parsed JSON), error, error_type
    """
    try:
        data = json.loads(json_string)
        return _success_response(data)
    except json.JSONDecodeError as e:
        logger.error(f"JSON decode error: {e}", exc_info=True)
        return _error_response(f"Invalid JSON: {str(e)}", 'JSONDecodeError')
    except Exception as e:
        logger.error(f"Error parsing JSON string: {e}", exc_info=True)
        return _error_response(str(e), type(e).__name__)


def parse_xml_string(xml_string: str) -> Dict[str, Any]:
    """
    Parse XML string
    
    Args:
        xml_string: XML string to parse
        
    Returns:
        Dict with success, result (parsed XML as dict), error, error_type
    """
    if not XML_AVAILABLE:
        return _error_response("XML support not available", 'ImportError')
    
    try:
        root = ET.fromstring(xml_string)
        
        def xml_to_dict(element):
            result = {}
            if element.text and element.text.strip():
                result['text'] = element.text.strip()
            result['tag'] = element.tag
            result['attrib'] = element.attrib
            children = [xml_to_dict(child) for child in element]
            if children:
                result['children'] = children
            return result
        
        data = xml_to_dict(root)
        return _success_response(data)
    except ET.ParseError as e:
        logger.error(f"XML parse error: {e}", exc_info=True)
        return _error_response(f"Invalid XML: {str(e)}", 'ParseError')
    except Exception as e:
        logger.error(f"Error parsing XML string: {e}", exc_info=True)
        return _error_response(str(e), type(e).__name__)


def validate_json(json_string: str) -> Dict[str, Any]:
    """
    Validate JSON
    
    Args:
        json_string: JSON string to validate
        
    Returns:
        Dict with success, result (validation result), error, error_type
    """
    try:
        json.loads(json_string)
        return _success_response({'valid': True})
    except json.JSONDecodeError as e:
        return _success_response({
            'valid': False,
            'error': str(e),
            'position': e.pos if hasattr(e, 'pos') else None,
        })


def validate_xml(xml_string: str) -> Dict[str, Any]:
    """
    Validate XML
    
    Args:
        xml_string: XML string to validate
        
    Returns:
        Dict with success, result (validation result), error, error_type
    """
    if not XML_AVAILABLE:
        return _error_response("XML support not available", 'ImportError')
    
    try:
        ET.fromstring(xml_string)
        return _success_response({'valid': True})
    except ET.ParseError as e:
        return _success_response({
            'valid': False,
            'error': str(e),
            'position': e.position if hasattr(e, 'position') else None,
        })

